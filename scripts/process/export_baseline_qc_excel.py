#!/usr/bin/env python3

from __future__ import annotations

import argparse
from pathlib import Path
from typing import cast

import pandas as pd
from openpyxl.styles import PatternFill


REPO_ROOT = Path(__file__).resolve().parents[2]
ANALYSIS_OUTPUT_ROOT = REPO_ROOT / "output" / "1.only_atac"


def is_bad(series: pd.Series, direction: str, nmads: float = 3.0) -> pd.Series:
    vals = cast(pd.Series, pd.to_numeric(series, errors="coerce"))
    med = vals.median(skipna=True)
    mad = (vals - med).abs().median(skipna=True)
    if pd.isna(mad) or mad == 0:
        return pd.Series(False, index=series.index)
    if direction == "low":
        return vals < (med - nmads * mad)
    return vals > (med + nmads * mad)


def main() -> None:
    parser = argparse.ArgumentParser(description="Export baseline QC summary to Excel")
    parser.add_argument(
        "--input-csv",
        default=str(
            ANALYSIS_OUTPUT_ROOT / "qc_reports" / "matrix_lite_qc_summary_baseline.csv"
        ),
    )
    parser.add_argument(
        "--output-xlsx",
        default=str(
            ANALYSIS_OUTPUT_ROOT
            / "qc_reports"
            / "matrix_lite_qc_summary_baseline_by_dataset.xlsx"
        ),
    )
    args = parser.parse_args()

    input_csv = Path(args.input_csv)
    output_xlsx = Path(args.output_xlsx)
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)

    df = pd.read_csv(input_csv)

    dataset_parts = []
    for gse, sub in df.groupby("gse", sort=False):
        sub = sub.copy()
        sub["dataset_flag_low_qc_rate"] = is_bad(cast(pd.Series, sub["qc_rate"]), "low")
        sub["dataset_flag_low_tss"] = is_bad(
            cast(pd.Series, sub["median_TSS_enrichment"]), "low"
        )
        sub["dataset_flag_low_frip"] = is_bad(
            cast(pd.Series, sub["median_FRiP"]), "low"
        )
        sub["dataset_flag_high_blacklist"] = is_bad(
            cast(pd.Series, sub["median_blacklist_fraction"]), "high"
        )
        sub["dataset_flag_high_doublet"] = is_bad(
            cast(pd.Series, sub["doublet_rate"]), "high"
        )
        sub["dataset_flag_any"] = sub[
            [
                "dataset_flag_low_qc_rate",
                "dataset_flag_low_tss",
                "dataset_flag_low_frip",
                "dataset_flag_high_blacklist",
                "dataset_flag_high_doublet",
            ]
        ].any(axis=1)
        dataset_parts.append(sub)

    df = pd.concat(dataset_parts, ignore_index=True)
    df["是否可以接受"] = [
        "否" if bool(flag) else "是" for flag in df["dataset_flag_any"].tolist()
    ]

    ordered_cols = [
        "gse",
        "gsm",
        "input_cells",
        "pass_qc",
        "qc_rate",
        "singlets",
        "doublets",
        "doublet_rate",
        "mean_nCount_ATAC",
        "median_fragments",
        "median_TSS_enrichment",
        "median_FRiP",
        "median_unique_ratio",
        "median_blacklist_fraction",
        "median_cima_l4_score",
        "cima_unique_l4_labels",
        "是否可以接受",
    ]
    ordered_cols = [c for c in ordered_cols if c in df.columns]
    out = df[ordered_cols].copy()

    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as writer:
        out.to_excel(writer, sheet_name="baseline_qc", index=False)
        ws = writer.book["baseline_qc"]
        ws.freeze_panes = "A2"

        accepted_fill = PatternFill(fill_type="solid", fgColor="EAF4E2")
        rejected_fill = PatternFill(fill_type="solid", fgColor="FCE8E6")
        low_metric_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
        high_metric_fill = PatternFill(fill_type="solid", fgColor="F4CCCC")

        col_index = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
        low_metric_cols = [
            "qc_rate",
            "median_TSS_enrichment",
            "median_FRiP",
        ]
        high_metric_cols = [
            "doublet_rate",
            "median_blacklist_fraction",
        ]

        metric_to_flag = {
            "qc_rate": "dataset_flag_low_qc_rate",
            "median_TSS_enrichment": "dataset_flag_low_tss",
            "median_FRiP": "dataset_flag_low_frip",
            "doublet_rate": "dataset_flag_high_doublet",
            "median_blacklist_fraction": "dataset_flag_high_blacklist",
        }

        for column_cells in ws.columns:
            max_len = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            ws.column_dimensions[column_cells[0].column_letter].width = min(
                max(max_len + 2, 12), 28
            )

        for row_idx, (_, src_row) in enumerate(df.iterrows(), start=2):
            row_fill = (
                accepted_fill if src_row["是否可以接受"] == "是" else rejected_fill
            )
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = row_fill

            for metric in low_metric_cols:
                if bool(src_row[metric_to_flag[metric]]):
                    ws.cell(
                        row=row_idx, column=col_index[metric]
                    ).fill = low_metric_fill
            for metric in high_metric_cols:
                if bool(src_row[metric_to_flag[metric]]):
                    ws.cell(
                        row=row_idx, column=col_index[metric]
                    ).fill = high_metric_fill

    print(output_xlsx)


if __name__ == "__main__":
    main()
