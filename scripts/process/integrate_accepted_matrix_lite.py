#!/usr/bin/env python3

from __future__ import annotations

import argparse
import gc
import threading
import time
from pathlib import Path

import anndata as ad
import matplotlib

matplotlib.use("Agg")

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import psutil
import scanpy as sc
import scanpy.external as sce
import scipy.io
import scipy.sparse as sp
import umap
from matplotlib.colors import to_hex, to_rgb
from openpyxl import load_workbook


REPO_ROOT = Path(__file__).resolve().parents[2]
OUTPUT_ROOT = REPO_ROOT / "output"
ANALYSIS_OUTPUT_ROOT = OUTPUT_ROOT / "1.only_atac"
DEFAULT_INPUT_ROOT = OUTPUT_ROOT
DEFAULT_REFERENCE_DIR = OUTPUT_ROOT / "reference" / "cima"
DEFAULT_DATASETS_XLSX = OUTPUT_ROOT / "reference" / "datasets.xlsx"

BASE_L1_PALETTE = {
    "B": "#2C7BB6",
    "CD4_T": "#D7191C",
    "CD8_T&unconvensional_T": "#FDAE61",
    "Myeloid": "#1A9641",
    "ILC": "#762A83",
}
FALLBACK_COLORS = [
    "#5E4FA2",
    "#3288BD",
    "#66C2A5",
    "#ABDDA4",
    "#FEE08B",
    "#F46D43",
    "#A50026",
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Integrate accepted matrix-lite samples"
    )
    parser.add_argument(
        "--input-root",
        default=str(DEFAULT_INPUT_ROOT),
        help="Root directory containing matrix-lite sample outputs",
    )
    parser.add_argument(
        "--accepted-xlsx",
        default=str(
            ANALYSIS_OUTPUT_ROOT
            / "qc_reports"
            / "matrix_lite_qc_summary_baseline_by_dataset.xlsx"
        ),
    )
    parser.add_argument(
        "--output-dir",
        default=str(ANALYSIS_OUTPUT_ROOT / "accepted_integration"),
    )
    parser.add_argument(
        "--reference-dir",
        default=str(DEFAULT_REFERENCE_DIR),
        help="Directory containing CIMA compact LSI assets",
    )
    parser.add_argument(
        "--datasets-xlsx",
        default=str(DEFAULT_DATASETS_XLSX),
    )
    parser.add_argument(
        "--method",
        choices=["pooled-umap", "bbknn"],
        default="pooled-umap",
        help="Integration method applied after pooling projected sample embeddings",
    )
    parser.add_argument(
        "--batch-key",
        choices=["gsm", "gse"],
        default="gsm",
        help="Batch covariate used by graph-based integration methods",
    )
    parser.add_argument(
        "--lsi-dims",
        type=int,
        default=30,
        help="Number of projected LSI dimensions to use after dropping dim_1",
    )
    parser.add_argument(
        "--neighbors-within-batch",
        type=int,
        default=1,
        help="BBKNN neighbors to keep from each batch",
    )
    parser.add_argument(
        "--trim",
        type=int,
        default=60,
        help="BBKNN trim value controlling graph density",
    )
    parser.add_argument(
        "--annoy-n-trees",
        type=int,
        default=20,
        help="BBKNN Annoy tree count for approximate neighbor search",
    )
    parser.add_argument(
        "--metric",
        default="euclidean",
        help="Distance metric for BBKNN",
    )
    parser.add_argument(
        "--umap-n-neighbors",
        type=int,
        default=30,
        help="UMAP neighbor count for pooled-umap mode",
    )
    parser.add_argument(
        "--umap-min-dist",
        type=float,
        default=0.3,
        help="UMAP minimum distance for visualization",
    )
    parser.add_argument(
        "--memory-log-interval-sec",
        type=float,
        default=2.0,
        help="Sampling interval for RAM/swap monitoring",
    )
    return parser.parse_args()


def mix_color(color: str, target: str, amount: float) -> str:
    amount = min(max(amount, 0.0), 1.0)
    source_rgb = np.asarray(to_rgb(color), dtype=np.float32)
    target_rgb = np.asarray(to_rgb(target), dtype=np.float32)
    blended = source_rgb * (1 - amount) + target_rgb * amount
    return to_hex(
        (float(blended[0]), float(blended[1]), float(blended[2])), keep_alpha=False
    )


def make_shade_palette(base_color: str, labels: list[str]) -> dict[str, str]:
    labels = list(dict.fromkeys(labels))
    if not labels:
        return {}
    if len(labels) == 1:
        return {labels[0]: base_color}
    colors = []
    for frac in np.linspace(0, 1, len(labels)):
        if frac <= 0.5:
            amount = 0.55 * (1 - frac / 0.5)
            colors.append(mix_color(base_color, "#FFFFFF", amount))
        else:
            amount = 0.18 * ((frac - 0.5) / 0.5)
            colors.append(mix_color(base_color, "#000000", amount))
    return dict(zip(labels, colors, strict=False))


def build_cima_palettes(hierarchy: pd.DataFrame) -> dict[str, dict[str, str]]:
    l1_labels = hierarchy["cell_type_l1"].drop_duplicates().tolist()
    l1_palette = dict(BASE_L1_PALETTE)
    missing = [label for label in l1_labels if label not in l1_palette]
    for label, color in zip(missing, FALLBACK_COLORS, strict=False):
        l1_palette[label] = color

    def level_palette(level_col: str) -> dict[str, str]:
        palette: dict[str, str] = {}
        for l1 in l1_labels:
            labels = sorted(
                hierarchy.loc[hierarchy["cell_type_l1"] == l1, level_col]
                .dropna()
                .unique()
                .tolist()
            )
            palette.update(make_shade_palette(l1_palette[l1], labels))
        return palette

    return {
        "cima_cell_type_l1": {label: l1_palette[label] for label in l1_labels},
        "cima_cell_type_l2": level_palette("cell_type_l2"),
        "cima_cell_type_l3": level_palette("cell_type_l3"),
        "cima_cell_type_l4": level_palette("cell_type_l4"),
    }


class MemoryMonitor:
    def __init__(self, interval_sec: float) -> None:
        self.interval_sec = interval_sec
        self.process = psutil.Process()
        self.rows: list[dict[str, float | str]] = []
        self._stop_event = threading.Event()
        self._thread = threading.Thread(target=self._run, daemon=True)
        self._start_time = time.time()

    def _snapshot(self, phase: str) -> dict[str, float | str]:
        vm = psutil.virtual_memory()
        sw = psutil.swap_memory()
        rss = self.process.memory_info().rss
        return {
            "elapsed_sec": round(time.time() - self._start_time, 3),
            "phase": phase,
            "rss_gib": round(rss / 2**30, 4),
            "available_gib": round(vm.available / 2**30, 4),
            "memory_percent": round(float(vm.percent), 2),
            "swap_used_gib": round(sw.used / 2**30, 4),
            "swap_percent": round(float(sw.percent), 2),
        }

    def _run(self) -> None:
        while not self._stop_event.wait(self.interval_sec):
            self.rows.append(self._snapshot("poll"))

    def start(self) -> None:
        self.rows.append(self._snapshot("start"))
        self._thread.start()

    def mark(self, phase: str) -> None:
        self.rows.append(self._snapshot(phase))

    def stop(self) -> None:
        self._stop_event.set()
        self._thread.join(timeout=self.interval_sec + 1.0)
        self.rows.append(self._snapshot("stop"))

    def write(self, path: Path) -> None:
        pd.DataFrame(self.rows).to_csv(path, index=False)

    def summary(self) -> dict[str, float]:
        memory_df = pd.DataFrame(self.rows)
        rss = memory_df["rss_gib"].to_numpy(dtype=np.float64)
        available = memory_df["available_gib"].to_numpy(dtype=np.float64)
        swap_used = memory_df["swap_used_gib"].to_numpy(dtype=np.float64)
        return {
            "peak_rss_gib": round(float(rss.max()), 4),
            "min_available_gib": round(float(available.min()), 4),
            "max_swap_used_gib": round(float(swap_used.max()), 4),
        }


def read_accepted_samples(workbook: Path) -> list[tuple[str, str]]:
    wb = load_workbook(workbook, read_only=True)
    ws = wb["baseline_qc"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    gse_i = header.index("gse")
    gsm_i = header.index("gsm")
    acc_i = header.index("是否可以接受")
    accepted = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[acc_i] == "是":
            accepted.append((str(row[gse_i]), str(row[gsm_i])))
    return accepted


def read_health_map(datasets_xlsx: Path) -> dict[str, str]:
    df = pd.read_excel(datasets_xlsx, sheet_name="pbmc_samples")
    df = df[["样本名(GSM*)", "健康状态"]].dropna()
    df["样本名(GSM*)"] = df["样本名(GSM*)"].astype(str)
    return dict(zip(df["样本名(GSM*)"], df["健康状态"], strict=False))


def load_feature_model(
    path: Path,
) -> tuple[np.ndarray, np.ndarray, np.ndarray, np.ndarray]:
    model_df = pd.read_csv(path, sep="\t")
    feature_index = model_df["feature_index"].astype(int).to_numpy() - 1
    feature_id = model_df["feature_id"].astype(str).to_numpy()
    dim_cols = [c for c in model_df.columns if c.startswith("dim_")]
    idf = model_df["idf"].to_numpy(dtype=np.float32)
    loadings = model_df[dim_cols].to_numpy(dtype=np.float32)
    return feature_index, feature_id, idf, loadings


def project_binary_counts(
    counts: sp.spmatrix, idf: np.ndarray, loadings: np.ndarray
) -> np.ndarray:
    selected = sp.csc_matrix(counts, dtype=np.float32, copy=True)
    if selected.nnz:
        selected.data[:] = 1.0
        col_totals = np.asarray(selected.sum(axis=0)).ravel().astype(np.float32)
        col_totals[col_totals == 0] = 1.0
        selected.data /= np.repeat(col_totals, np.diff(selected.indptr))
    tfidf = sp.diags(idf) @ selected
    return np.asarray(tfidf.T @ loadings, dtype=np.float32)


def load_sample(
    sample_dir: Path,
    feature_index: np.ndarray,
    feature_id: np.ndarray,
    idf: np.ndarray,
    loadings: np.ndarray,
) -> tuple[pd.DataFrame, np.ndarray]:
    meta = pd.read_csv(sample_dir / "validation_result.csv")
    features = (
        pd.read_csv(sample_dir / "matrix" / "features.tsv", header=None)[0]
        .astype(str)
        .to_numpy()
    )
    observed = features[feature_index]
    if not np.array_equal(observed, feature_id):
        mismatch = int(np.where(observed != feature_id)[0][0])
        raise ValueError(
            f"Feature mismatch for {sample_dir.name} at index {mismatch}: {observed[mismatch]} != {feature_id[mismatch]}"
        )
    matrix = sp.csr_matrix(scipy.io.mmread(sample_dir / "matrix" / "matrix.mtx"))
    selected = matrix[feature_index, :]
    embedding = project_binary_counts(selected, idf, loadings)
    return meta, embedding


def prepare_integration_representation(
    integrated_embedding: np.ndarray, lsi_dims: int
) -> np.ndarray:
    if integrated_embedding.shape[1] <= 1:
        raise ValueError("Projected embedding must contain at least two LSI dimensions")
    n_dims = min(lsi_dims, integrated_embedding.shape[1] - 1)
    if n_dims <= 0:
        raise ValueError(f"Invalid --lsi-dims value: {lsi_dims}")
    return np.asarray(integrated_embedding[:, 1 : 1 + n_dims], dtype=np.float32)


def optimize_metadata(df: pd.DataFrame) -> pd.DataFrame:
    optimized = df.copy()
    categorical_columns = [
        "gse",
        "gsm",
        "sample",
        "health_status",
        "scDblFinder.class",
        "cima_cell_type_l1",
        "cima_cell_type_l2",
        "cima_cell_type_l3",
        "cima_cell_type_l4",
    ]
    for column in categorical_columns:
        if column in optimized.columns:
            optimized[column] = optimized[column].astype("category")
    return optimized


def run_pooled_umap(rep: np.ndarray, args: argparse.Namespace) -> np.ndarray:
    reducer = umap.UMAP(
        n_neighbors=args.umap_n_neighbors,
        min_dist=args.umap_min_dist,
        metric="cosine",
        random_state=42,
        transform_seed=42,
        low_memory=True,
    )
    return np.asarray(reducer.fit_transform(rep), dtype=np.float32)


def run_bbknn_umap(
    meta: pd.DataFrame, rep: np.ndarray, args: argparse.Namespace
) -> np.ndarray:
    obs = meta.copy()
    obs.index = obs.index.astype(str)
    adata = ad.AnnData(obs=obs)
    adata.obsm["X_lsi"] = np.asarray(rep, dtype=np.float32)
    sce.pp.bbknn(
        adata,
        batch_key=args.batch_key,
        use_rep="X_lsi",
        n_pcs=adata.obsm["X_lsi"].shape[1],
        neighbors_within_batch=args.neighbors_within_batch,
        trim=args.trim,
        approx=True,
        use_annoy=True,
        annoy_n_trees=args.annoy_n_trees,
        metric=args.metric,
        computation="annoy",
    )
    sc.tl.umap(adata, random_state=42, min_dist=args.umap_min_dist)
    return np.asarray(adata.obsm["X_umap"], dtype=np.float32)


def plot_categorical(
    df: pd.DataFrame,
    x: str,
    y: str,
    color_col: str,
    palette: dict[str, str],
    out_path: Path,
    title: str,
) -> None:
    fig, ax = plt.subplots(figsize=(10, 8), constrained_layout=True)
    ordered = [
        label for label in palette if label in set(df[color_col].dropna().astype(str))
    ]
    for label in ordered:
        sub = df[df[color_col].astype(str) == label]
        ax.scatter(
            sub[x], sub[y], s=3, alpha=0.75, c=palette[label], label=label, linewidths=0
        )
    ax.set_title(title, fontweight="bold")
    ax.set_xlabel(x)
    ax.set_ylabel(y)
    ax.legend(
        loc="upper left",
        bbox_to_anchor=(1.02, 1),
        frameon=False,
        markerscale=3,
        fontsize=8,
    )
    fig.savefig(out_path, dpi=150)
    plt.close(fig)


def plot_numeric(
    df: pd.DataFrame, x: str, y: str, color_col: str, out_path: Path, title: str
) -> None:
    fig, ax = plt.subplots(figsize=(10, 8), constrained_layout=True)
    sc = ax.scatter(
        df[x], df[y], s=3, alpha=0.75, c=df[color_col], cmap="viridis", linewidths=0
    )
    ax.set_title(title, fontweight="bold")
    ax.set_xlabel(x)
    ax.set_ylabel(y)
    cbar = fig.colorbar(sc, ax=ax)
    cbar.set_label(color_col)
    fig.savefig(out_path, dpi=150)
    plt.close(fig)


def main() -> None:
    args = parse_args()
    input_root = Path(args.input_root)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    reference_dir = Path(args.reference_dir)
    memory_log_path = output_dir / "accepted_integration_memory_log.csv"
    memory_monitor = MemoryMonitor(args.memory_log_interval_sec)
    memory_monitor.start()

    try:
        accepted = read_accepted_samples(Path(args.accepted_xlsx))
        health_map = read_health_map(Path(args.datasets_xlsx))
        feature_index, feature_id, idf, loadings = load_feature_model(
            reference_dir / "cima_atac_reference_lsi_features.tsv.gz"
        )
        hierarchy = pd.read_csv(reference_dir / "cima_atac_celltype_hierarchy.csv")
        palettes = build_cima_palettes(hierarchy)
        memory_monitor.mark("reference_loaded")

        meta_parts: list[pd.DataFrame] = []
        embed_parts: list[np.ndarray] = []
        for index, (gse, gsm) in enumerate(accepted, start=1):
            sample_dir = input_root / gse / gsm
            meta, embedding = load_sample(
                sample_dir, feature_index, feature_id, idf, loadings
            )
            meta = meta.copy()
            meta["gse"] = gse
            meta["gsm"] = gsm
            meta["sample"] = f"{gse}/{gsm}"
            meta["health_status"] = health_map.get(gsm, "未知")
            meta_parts.append(meta)
            embed_parts.append(embedding)
            if index % 10 == 0 or index == len(accepted):
                memory_monitor.mark(f"samples_loaded_{index}")

        integrated_meta = optimize_metadata(pd.concat(meta_parts, ignore_index=True))
        integrated_embedding = np.vstack(embed_parts).astype(np.float32)
        del meta_parts
        del embed_parts
        gc.collect()
        memory_monitor.mark("embeddings_stacked")

        integration_rep = prepare_integration_representation(
            integrated_embedding, args.lsi_dims
        )
        del integrated_embedding
        gc.collect()
        memory_monitor.mark("integration_rep_prepared")

        if args.method == "bbknn":
            integrated_umap = run_bbknn_umap(integrated_meta, integration_rep, args)
        else:
            integrated_umap = run_pooled_umap(integration_rep, args)
        del integration_rep
        gc.collect()
        memory_monitor.mark(f"{args.method}_completed")

        integrated_meta["integrated_umap_1"] = integrated_umap[:, 0]
        integrated_meta["integrated_umap_2"] = integrated_umap[:, 1]
        integrated_meta.to_csv(
            output_dir / "accepted_integration_metadata.csv", index=False
        )
        memory_monitor.mark("metadata_written")

        gse_palette = {
            g: c
            for g, c in zip(
                sorted(integrated_meta["gse"].unique()),
                ["#1f78b4", "#e31a1c"],
                strict=False,
            )
        }
        gsm_values = sorted(integrated_meta["gsm"].unique())
        cmap = plt.get_cmap("tab20", len(gsm_values))
        gsm_palette = {gsm: to_hex(cmap(i)) for i, gsm in enumerate(gsm_values)}
        health_values = sorted(integrated_meta["health_status"].unique())
        health_palette_base = {
            "Healthy": "#2C7BB6",
            "COVID-19": "#D7191C",
            "RSV": "#1A9641",
            "Unknown": "#7F7F7F",
        }
        health_label_map = {
            "健康": "Healthy",
            "COVID-19": "COVID-19",
            "RSV": "RSV",
            "未知": "Unknown",
        }
        integrated_meta["health_status_plot"] = (
            integrated_meta["health_status"]
            .astype(str)
            .map(lambda value: health_label_map.get(value, value))
            .fillna(integrated_meta["health_status"].astype(str))
        )
        health_palette = {
            health: health_palette_base.get(health, "#7F7F7F")
            for health in sorted(integrated_meta["health_status_plot"].unique())
        }

        for level, title_suffix in [
            ("cima_cell_type_l1", "CIMA L1"),
            ("cima_cell_type_l2", "CIMA L2"),
            ("cima_cell_type_l3", "CIMA L3"),
            ("cima_cell_type_l4", "CIMA L4"),
        ]:
            plot_categorical(
                integrated_meta,
                "integrated_umap_1",
                "integrated_umap_2",
                level,
                palettes[level],
                output_dir
                / f"accepted_integration_{level.replace('cima_cell_type_', 'cima_')}.png",
                f"Accepted-sample integration by {title_suffix}",
            )
        plot_categorical(
            integrated_meta,
            "integrated_umap_1",
            "integrated_umap_2",
            "gse",
            gse_palette,
            output_dir / "accepted_integration_gse.png",
            "Accepted-sample integration by GSE",
        )
        plot_categorical(
            integrated_meta,
            "integrated_umap_1",
            "integrated_umap_2",
            "gsm",
            gsm_palette,
            output_dir / "accepted_integration_gsm.png",
            "Accepted-sample integration by GSM",
        )
        plot_categorical(
            integrated_meta,
            "integrated_umap_1",
            "integrated_umap_2",
            "health_status_plot",
            health_palette,
            output_dir / "accepted_integration_health_status.png",
            "Accepted-sample integration by health status",
        )

        for metric in ["nCount_ATAC", "TSS.enrichment", "FRiP", "blacklist_fraction"]:
            plot_numeric(
                integrated_meta,
                "integrated_umap_1",
                "integrated_umap_2",
                metric,
                output_dir / f"accepted_integration_{metric.replace('.', '_')}.png",
                f"Accepted-sample integration by {metric}",
            )
        memory_monitor.mark("plots_written")
    except Exception:
        memory_monitor.mark("error")
        raise
    finally:
        memory_monitor.stop()
        memory_monitor.write(memory_log_path)

    summary_row = {
        "integration_method": args.method,
        "batch_key": args.batch_key if args.method == "bbknn" else "none",
        "accepted_samples": len(accepted),
        "accepted_cells": len(integrated_meta),
        "gse_count": len(set(map(str, integrated_meta["gse"].tolist()))),
        "health_status_levels": len(
            set(map(str, integrated_meta["health_status"].tolist()))
        ),
        "integration_dims": args.lsi_dims,
        "neighbors_within_batch": args.neighbors_within_batch
        if args.method == "bbknn"
        else np.nan,
        "trim": args.trim if args.method == "bbknn" else np.nan,
        "annoy_n_trees": args.annoy_n_trees if args.method == "bbknn" else np.nan,
        **memory_monitor.summary(),
    }
    pd.DataFrame([summary_row]).to_csv(
        output_dir / "accepted_integration_summary.csv", index=False
    )

    print(output_dir)


if __name__ == "__main__":
    main()
